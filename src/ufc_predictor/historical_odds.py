from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .io import ensure_parent, normalize_name
from .odds import rank_value_bets


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY_FMT = "$#,##0.00;[Red]-$#,##0.00"
PCT_FMT = "0.0%"
NUM_FMT = "0.000"


def backtest_historical_odds(
    predictions: pd.DataFrame,
    historical_odds: pd.DataFrame,
    bankroll: float = 1000.0,
    flat_stake: float = 10.0,
    max_confidence_stake: float = 100.0,
    kelly_multiplier: float = 0.25,
    max_bankroll_fraction: float = 0.02,
    min_edge: float = 0.02,
    min_expected_roi: float = 0.0,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    _validate_prediction_results(predictions)
    ranked = rank_value_bets(
        predictions,
        historical_odds,
        bankroll=bankroll,
        kelly_multiplier=kelly_multiplier,
        max_bankroll_fraction=max_bankroll_fraction,
        min_edge=min_edge,
        min_expected_roi=min_expected_roi,
    )
    actual_lookup = _actual_winner_lookup(predictions)
    rows: list[dict[str, object]] = []
    for raw in ranked.to_dict("records"):
        key = _matchup_key(raw)
        actual_winner = actual_lookup[key]
        is_bet = raw["decision"] == "bet"
        won = normalize_name(raw["fighter"]) == normalize_name(actual_winner)
        decimal_odds = float(raw["decimal_odds"])
        model_probability = float(raw["model_probability"])
        confidence = max(0.0, min(1.0, (model_probability - 0.5) * 2.0))
        kelly_stake = float(raw["recommended_stake"]) if is_bet else 0.0
        flat_bet_stake = flat_stake if is_bet else 0.0
        confidence_stake = max_confidence_stake * confidence if is_bet else 0.0
        output = dict(raw)
        output.update(
            {
                "actual_winner": actual_winner,
                "bet_won": bool(won and is_bet),
                "bet_placed": bool(is_bet),
                "confidence": confidence,
                "kelly_stake": kelly_stake,
                "kelly_profit": _profit(won, kelly_stake, decimal_odds),
                "flat_stake": flat_bet_stake,
                "flat_profit": _profit(won, flat_bet_stake, decimal_odds),
                "confidence_stake": confidence_stake,
                "confidence_profit": _profit(won, confidence_stake, decimal_odds),
            }
        )
        rows.append(output)

    backtest = pd.DataFrame(rows)
    if backtest.empty:
        return backtest, _empty_strategy_summary()
    backtest = backtest.sort_values(["event_date", "fighter_a", "fighter_b", "sportsbook", "fighter"]).reset_index(drop=True)
    _add_cumulative_columns(backtest, bankroll, "kelly")
    _add_cumulative_columns(backtest, bankroll, "flat")
    _add_cumulative_columns(backtest, bankroll, "confidence")
    return backtest, summarize_strategies(backtest)


def summarize_strategies(backtest: pd.DataFrame) -> pd.DataFrame:
    if backtest.empty:
        return _empty_strategy_summary()
    rows: list[dict[str, object]] = []
    for strategy in ["kelly", "flat", "confidence"]:
        stake_col = f"{strategy}_stake"
        profit_col = f"{strategy}_profit"
        bankroll_col = f"{strategy}_bankroll"
        drawdown_col = f"{strategy}_drawdown"
        placed = backtest[backtest[stake_col] > 0].copy()
        total_staked = float(placed[stake_col].sum())
        net_profit = float(placed[profit_col].sum())
        rows.append(
            {
                "strategy": strategy,
                "bets": int(len(placed)),
                "wins": int(placed["bet_won"].sum()) if not placed.empty else 0,
                "hit_rate": float(placed["bet_won"].mean()) if not placed.empty else None,
                "total_staked": total_staked,
                "net_profit": net_profit,
                "roi": net_profit / total_staked if total_staked else None,
                "ending_bankroll": float(backtest[bankroll_col].iloc[-1]),
                "max_drawdown": float(backtest[drawdown_col].min()),
            }
        )
    return pd.DataFrame(rows)


def write_historical_odds_workbook(
    output_path: Path,
    backtest: pd.DataFrame,
    strategy_summary: pd.DataFrame,
) -> None:
    ensure_parent(output_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_readme_sheet(workbook)
    _write_dataframe_sheet(workbook, "Strategy Summary", strategy_summary, "StrategySummaryTable")
    _write_dataframe_sheet(workbook, "Bet Backtest", backtest, "BetBacktestTable")
    _write_chart_data_and_charts(workbook, backtest, strategy_summary)
    workbook.save(output_path)


def _validate_prediction_results(predictions: pd.DataFrame) -> None:
    required = {"event_date", "fighter_a", "fighter_b", "prob_fighter_a", "prob_fighter_b"}
    has_actual_name = "actual_winner" in predictions.columns
    has_target = "target_fighter_a_win" in predictions.columns
    missing = sorted(required - set(predictions.columns))
    if missing:
        raise ValueError(f"historical predictions missing columns: {missing}")
    if not has_actual_name and not has_target:
        raise ValueError("historical predictions must include actual_winner or target_fighter_a_win.")
    if has_actual_name and not has_target and predictions["actual_winner"].isna().any():
        raise ValueError("actual_winner cannot be missing unless target_fighter_a_win is also provided.")


def _actual_winner_lookup(predictions: pd.DataFrame) -> dict[tuple[str, str, str], str]:
    lookup: dict[tuple[str, str, str], str] = {}
    for row in predictions.to_dict("records"):
        if "actual_winner" in row and pd.notna(row["actual_winner"]):
            actual = str(row["actual_winner"])
        else:
            actual = str(row["fighter_a"] if int(row["target_fighter_a_win"]) == 1 else row["fighter_b"])
        lookup[_matchup_key(row)] = actual
    return lookup


def _matchup_key(row: dict[str, object]) -> tuple[str, str, str]:
    fighters = sorted([normalize_name(row["fighter_a"]), normalize_name(row["fighter_b"])])
    return str(row["event_date"]), fighters[0], fighters[1]


def _profit(won: bool, stake: float, decimal_odds: float) -> float:
    if stake <= 0:
        return 0.0
    return stake * (decimal_odds - 1.0) if won else -stake


def _add_cumulative_columns(backtest: pd.DataFrame, bankroll: float, strategy: str) -> None:
    profit_col = f"{strategy}_profit"
    cumulative_col = f"{strategy}_cumulative_profit"
    bankroll_col = f"{strategy}_bankroll"
    drawdown_col = f"{strategy}_drawdown"
    backtest[cumulative_col] = backtest[profit_col].cumsum()
    backtest[bankroll_col] = bankroll + backtest[cumulative_col]
    running_max = backtest[bankroll_col].cummax()
    backtest[drawdown_col] = backtest[bankroll_col] - running_max


def _empty_strategy_summary() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "strategy",
            "bets",
            "wins",
            "hit_rate",
            "total_staked",
            "net_profit",
            "roi",
            "ending_bankroll",
            "max_drawdown",
        ]
    )


def _write_readme_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("README")
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "Historical Odds Backtest"
    sheet["A1"].font = Font(bold=True, size=16)
    rows = [
        ("What this is", "A replay of model-ranked bets against historical sportsbook moneylines."),
        ("Required inputs", "Historical model predictions with actual winners plus a historical odds board."),
        ("Strategies", "Conservative Kelly, flat stake, and confidence stake are compared side by side."),
        ("Pass discipline", "Rows that fail the configured edge/EV filters are preserved with zero stake."),
        ("Risk framing", "Positive expected value can still lose and drawdowns can be large."),
    ]
    for row_idx, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row_idx, column=2, value=value)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 110


def _write_dataframe_sheet(workbook: Workbook, sheet_name: str, frame: pd.DataFrame, table_name: str) -> None:
    sheet = workbook.create_sheet(sheet_name)
    safe_frame = frame.copy()
    if safe_frame.empty:
        if len(safe_frame.columns) == 0:
            safe_frame = pd.DataFrame({"message": ["No rows available."]})
        else:
            safe_frame = pd.DataFrame([{column: None for column in safe_frame.columns}])
    for col_idx, column in enumerate(safe_frame.columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=str(column))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="top")
    for row_idx, row in enumerate(safe_frame.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=_excel_value(value))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    table_ref = f"A1:{get_column_letter(len(safe_frame.columns))}{len(safe_frame) + 1}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    sheet.add_table(table)
    _format_columns(sheet, safe_frame)


def _write_chart_data_and_charts(workbook: Workbook, backtest: pd.DataFrame, strategy_summary: pd.DataFrame) -> None:
    chart_data = workbook.create_sheet("Chart Data")
    chart_data.append(["bet_number", "kelly_bankroll", "flat_bankroll", "confidence_bankroll"])
    if not backtest.empty:
        for idx, row in enumerate(backtest.itertuples(index=False), start=1):
            chart_data.append([idx, row.kelly_bankroll, row.flat_bankroll, row.confidence_bankroll])
    summary_start = max(len(backtest), 1) + 4
    chart_data.cell(row=summary_start, column=1, value="strategy")
    chart_data.cell(row=summary_start, column=2, value="net_profit")
    chart_data.cell(row=summary_start, column=3, value="max_drawdown")
    for offset, row in enumerate(strategy_summary.itertuples(index=False), start=1):
        chart_data.cell(row=summary_start + offset, column=1, value=row.strategy)
        chart_data.cell(row=summary_start + offset, column=2, value=row.net_profit)
        chart_data.cell(row=summary_start + offset, column=3, value=row.max_drawdown)
    for row_idx in [1, summary_start]:
        for col_idx in range(1, 4):
            cell = chart_data.cell(row=row_idx, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
    chart_data.sheet_state = "hidden"

    charts = workbook.create_sheet("Charts")
    charts.sheet_view.showGridLines = False
    charts["A1"] = "Historical Odds Backtest Charts"
    charts["A1"].font = Font(bold=True, size=16)
    if len(backtest) > 0:
        line = LineChart()
        line.title = "Bankroll By Strategy"
        line.y_axis.title = "Bankroll ($)"
        line.x_axis.title = "Ranked Bet Number"
        line.add_data(Reference(chart_data, min_col=2, max_col=4, min_row=1, max_row=len(backtest) + 1), titles_from_data=True)
        line.set_categories(Reference(chart_data, min_col=1, min_row=2, max_row=len(backtest) + 1))
        line.height = 9
        line.width = 24
        charts.add_chart(line, "A4")
    if len(strategy_summary) > 0:
        _add_bar_chart(charts, chart_data, "Net Profit By Strategy", "Net Profit ($)", 2, summary_start, summary_start + len(strategy_summary), "A22")
        _add_bar_chart(charts, chart_data, "Max Drawdown By Strategy", "Drawdown ($)", 3, summary_start, summary_start + len(strategy_summary), "J22")


def _add_bar_chart(sheet, data_sheet, title: str, y_title: str, data_col: int, header_row: int, end_row: int, anchor: str) -> None:
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = "Strategy"
    chart.add_data(Reference(data_sheet, min_col=data_col, min_row=header_row, max_row=end_row), titles_from_data=True)
    chart.set_categories(Reference(data_sheet, min_col=1, min_row=header_row + 1, max_row=end_row))
    chart.height = 9
    chart.width = 18
    sheet.add_chart(chart, anchor)


def _excel_value(value: object) -> object:
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if pd.isna(value):
        return None
    return value


def _format_columns(sheet, frame: pd.DataFrame) -> None:
    for col_idx, column in enumerate(frame.columns, start=1):
        sample = frame[column].head(250).astype(object)
        values = [str(column)] + [str(value) if pd.notna(value) else "" for value in sample.tolist()]
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(len(value) for value in values) + 2, 44)
        lower = str(column).lower()
        if any(token in lower for token in ["profit", "stake", "bankroll", "loss", "payout", "staked", "drawdown"]):
            _format_column(sheet, col_idx, len(frame), MONEY_FMT)
        elif any(token in lower for token in ["probability", "confidence", "edge", "roi", "fraction", "hit_rate"]):
            _format_column(sheet, col_idx, len(frame), PCT_FMT)
        elif any(token in lower for token in ["score", "volatility", "decimal_odds"]):
            _format_column(sheet, col_idx, len(frame), NUM_FMT)


def _format_column(sheet, col_idx: int, row_count: int, number_format: str) -> None:
    for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=row_count + 1):
        row[0].number_format = number_format
