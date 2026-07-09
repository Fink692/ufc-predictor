from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .evaluation import classification_metrics
from .features import MODEL_FEATURES
from .io import ensure_parent
from .models import _build_primary_pipeline, make_order_balanced_frame


CONFIDENCE_BUCKETS = [
    (0, 5, "0-5%"),
    (5, 10, "5-10%"),
    (10, 15, "10-15%"),
    (15, 20, "15-20%"),
    (20, 30, "20-30%"),
    (30, 40, "30-40%"),
    (40, 50, "40-50%"),
    (50, 100, "50-100%"),
]


def generate_backtest_workbook(
    features_path: Path,
    output_path: Path,
    rows_output_path: Path | None = None,
    max_confidence_stake: float = 100.0,
    starting_bankroll: float = 1000.0,
    holdout_fraction: float = 0.2,
    random_state: int = 42,
) -> dict[str, object]:
    if max_confidence_stake < 0:
        raise ValueError("max_confidence_stake cannot be negative.")
    if starting_bankroll < 0:
        raise ValueError("starting_bankroll cannot be negative.")
    if not 0 < holdout_fraction < 1:
        raise ValueError("holdout_fraction must be between 0 and 1.")

    features = pd.read_csv(features_path)
    backtest = build_backtest_frame(
        features,
        max_confidence_stake=max_confidence_stake,
        starting_bankroll=starting_bankroll,
        holdout_fraction=holdout_fraction,
        random_state=random_state,
    )
    summary, confidence, yearly, thresholds = summarize_backtest(
        backtest,
        source_rows=len(features),
        train_rows=len(features) - len(backtest),
        max_confidence_stake=max_confidence_stake,
        starting_bankroll=starting_bankroll,
    )

    ensure_parent(output_path)
    write_workbook(output_path, backtest, summary, confidence, yearly, thresholds)
    if rows_output_path is not None:
        ensure_parent(rows_output_path)
        backtest.to_csv(rows_output_path, index=False)

    total_staked = float(backtest["stake_if_100_confidence_equals_100"].sum())
    net_profit = float(backtest["net_profit_even_money_sim"].sum())
    return {
        "output_path": str(output_path),
        "rows_output_path": str(rows_output_path) if rows_output_path is not None else None,
        "holdout_rows": int(len(backtest)),
        "accuracy": float(backtest["correct_pick"].mean()) if len(backtest) else None,
        "total_staked_even_money": total_staked,
        "net_profit_even_money": net_profit,
        "roi_even_money": _roi(net_profit, total_staked),
        "max_drawdown_even_money": _max_drawdown(backtest["cumulative_profit_even_money_sim"]),
    }


def build_backtest_frame(
    features: pd.DataFrame,
    max_confidence_stake: float,
    starting_bankroll: float,
    holdout_fraction: float,
    random_state: int,
) -> pd.DataFrame:
    _validate_feature_columns(features)
    training = features.sort_values("event_date").reset_index(drop=True)
    test_size = max(1, int(round(len(training) * holdout_fraction)))
    train = training.iloc[:-test_size].copy()
    holdout = training.iloc[-test_size:].copy().reset_index(drop=True)
    if len(train) < 2:
        raise ValueError("Not enough training rows before the holdout split.")
    y_train = train["target_fighter_a_win"].astype(int)
    if len(set(y_train.tolist())) < 2 or int(y_train.value_counts().min()) < 2:
        raise ValueError("Training split must contain at least two examples of both classes.")

    balanced_train = make_order_balanced_frame(train)
    model = _build_primary_pipeline(random_state=random_state)
    model.fit(balanced_train[MODEL_FEATURES], balanced_train["target_fighter_a_win"].astype(int))
    prob_a = model.predict_proba(holdout[MODEL_FEATURES])[:, 1]
    prob_b = 1.0 - prob_a

    target = holdout["target_fighter_a_win"].astype(int).to_numpy()
    pick_a = prob_a >= 0.5
    pick_probability = np.where(pick_a, prob_a, prob_b)
    fighter_a = _fighter_column(holdout, "fighter_a")
    fighter_b = _fighter_column(holdout, "fighter_b")
    predicted_winner = np.where(pick_a, fighter_a, fighter_b)
    opponent = np.where(pick_a, fighter_b, fighter_a)
    actual_winner = np.where(target == 1, fighter_a, fighter_b)
    correct = predicted_winner == actual_winner
    confidence = np.clip((pick_probability - 0.5) * 2.0, 0.0, 1.0)
    stake = confidence * max_confidence_stake
    net_profit = np.where(correct, stake, -stake)

    backtest = pd.DataFrame(
        {
            "fight_id": holdout["fight_id"],
            "event_date": holdout["event_date"],
            "year": pd.to_datetime(holdout["event_date"]).dt.year,
            "fighter_a": fighter_a,
            "fighter_b": fighter_b,
            "weight_class": holdout["weight_class"],
            "gender": holdout["gender"],
            "scheduled_rounds": holdout["scheduled_rounds"],
            "title_fight": holdout["title_fight"],
            "actual_winner": actual_winner,
            "predicted_winner": predicted_winner,
            "opponent": opponent,
            "correct_pick": correct,
            "prob_fighter_a": prob_a,
            "prob_fighter_b": prob_b,
            "pick_win_probability": pick_probability,
            "confidence": confidence,
            "confidence_percent": confidence * 100.0,
            "confidence_bucket": [_confidence_bucket(value) for value in confidence * 100.0],
            "model_fair_odds_fighter_a": [_fair_american(value) for value in prob_a],
            "model_fair_odds_fighter_b": [_fair_american(value) for value in prob_b],
            "model_fair_odds_pick": [_fair_american(value) for value in pick_probability],
            "stake_if_100_confidence_equals_100": stake,
            "profit_at_even_money_if_correct": np.where(correct, stake, 0.0),
            "loss_at_even_money_if_wrong": np.where(correct, 0.0, stake),
            "net_profit_even_money_sim": net_profit,
            "expected_profit_even_money_sim": stake * ((pick_probability * 2.0) - 1.0),
            "cumulative_profit_even_money_sim": np.cumsum(net_profit),
            "bankroll_even_money_sim": starting_bankroll + np.cumsum(net_profit),
            "data_quality_flags": holdout.get("data_quality_flags", ""),
        }
    )
    backtest["event_date"] = pd.to_datetime(backtest["event_date"]).dt.date.astype(str)
    return backtest


def summarize_backtest(
    backtest: pd.DataFrame,
    source_rows: int,
    train_rows: int,
    max_confidence_stake: float,
    starting_bankroll: float,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    y_true = backtest["actual_winner"] == backtest["fighter_a"]
    metrics = classification_metrics(y_true.astype(int), backtest["prob_fighter_a"])
    total_staked = float(backtest["stake_if_100_confidence_equals_100"].sum())
    net_profit = float(backtest["net_profit_even_money_sim"].sum())
    summary = pd.DataFrame(
        [
            ("Dataset", "Historical chronological holdout"),
            ("Source feature rows", source_rows),
            ("Training rows used", train_rows),
            ("Holdout rows / simulated fights", len(backtest)),
            ("Holdout date range", f"{backtest['event_date'].min()} to {backtest['event_date'].max()}"),
            ("Model", "order_balanced_stamina_logistic"),
            ("Historical sportsbook odds available?", "No"),
            ("Odds shown in fight sheet", "Model-implied fair odds, not sportsbook lines"),
            ("P&L simulation assumption", "Even-money +100 payout using confidence-sized stakes"),
            ("Max confidence stake", max_confidence_stake),
            ("Starting bankroll shown", starting_bankroll),
            ("Accuracy", metrics["accuracy"]),
            ("Log loss", metrics["log_loss"]),
            ("Brier score", metrics["brier_score"]),
            ("ROC AUC", metrics["roc_auc"]),
            ("Correct picks", int(backtest["correct_pick"].sum())),
            ("Wrong picks", int((~backtest["correct_pick"]).sum())),
            ("Total simulated stake", total_staked),
            ("Net profit @ even money", net_profit),
            ("ROI @ even money", _roi(net_profit, total_staked)),
            ("Ending bankroll @ even money", float(backtest["bankroll_even_money_sim"].iloc[-1])),
            ("Max drawdown @ even money", _max_drawdown(backtest["cumulative_profit_even_money_sim"])),
            ("Average confidence", float(backtest["confidence"].mean())),
            ("Average stake", float(backtest["stake_if_100_confidence_equals_100"].mean())),
        ],
        columns=["Metric", "Value"],
    )

    confidence = _confidence_summary(backtest)
    yearly = _yearly_summary(backtest)
    thresholds = _threshold_summary(backtest)
    return summary, confidence, yearly, thresholds


def write_workbook(
    output_path: Path,
    backtest: pd.DataFrame,
    summary: pd.DataFrame,
    confidence: pd.DataFrame,
    yearly: pd.DataFrame,
    thresholds: pd.DataFrame,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_readme_sheet(workbook)
    _write_dataframe_sheet(workbook, "Summary", summary, "SummaryTable")
    _write_dataframe_sheet(workbook, "Fight Backtest", backtest, "FightBacktestTable")
    _write_dataframe_sheet(workbook, "Confidence Buckets", confidence, "ConfidenceBucketsTable")
    _write_dataframe_sheet(workbook, "Yearly", yearly, "YearlyTable")
    _write_dataframe_sheet(workbook, "Thresholds", thresholds, "ThresholdsTable")
    _write_chart_data_and_charts(workbook, backtest, confidence, yearly)
    workbook.save(output_path)


def _validate_feature_columns(features: pd.DataFrame) -> None:
    required = {"fight_id", "event_date", "target_fighter_a_win", "weight_class", "gender", "scheduled_rounds", "title_fight"}
    required.update(MODEL_FEATURES)
    if "fighter_a_name" not in features.columns and "fighter_a" not in features.columns:
        required.add("fighter_a_name")
    if "fighter_b_name" not in features.columns and "fighter_b" not in features.columns:
        required.add("fighter_b_name")
    missing = sorted(required - set(features.columns))
    if missing:
        raise ValueError(f"Feature table missing required columns: {missing}")


def _fighter_column(frame: pd.DataFrame, side: str) -> pd.Series:
    name_column = f"{side}_name"
    return frame[name_column] if name_column in frame.columns else frame[side]


def _fair_american(probability: float) -> int:
    value = min(max(float(probability), 1e-6), 1.0 - 1e-6)
    if value >= 0.5:
        return int(round(-100.0 * value / (1.0 - value)))
    return int(round(100.0 * (1.0 - value) / value))


def _confidence_bucket(confidence_percent: float) -> str:
    for low, high, label in CONFIDENCE_BUCKETS:
        if low <= confidence_percent < high:
            return label
    return CONFIDENCE_BUCKETS[-1][2]


def _roi(net_profit: float, stake: float) -> float:
    return float(net_profit / stake) if stake else 0.0


def _max_drawdown(series: pd.Series) -> float:
    running_max = series.cummax()
    return float((series - running_max).min()) if len(series) else 0.0


def _confidence_summary(backtest: pd.DataFrame) -> pd.DataFrame:
    summary = (
        backtest.groupby("confidence_bucket")
        .agg(
            fights=("fight_id", "count"),
            correct=("correct_pick", "sum"),
            avg_pick_probability=("pick_win_probability", "mean"),
            avg_confidence=("confidence", "mean"),
            total_staked=("stake_if_100_confidence_equals_100", "sum"),
            net_profit_even_money=("net_profit_even_money_sim", "sum"),
            expected_profit_even_money=("expected_profit_even_money_sim", "sum"),
        )
        .reset_index()
    )
    order = [label for *_bounds, label in CONFIDENCE_BUCKETS]
    summary["confidence_bucket"] = pd.Categorical(summary["confidence_bucket"], categories=order, ordered=True)
    summary = summary.sort_values("confidence_bucket")
    summary["accuracy"] = summary["correct"] / summary["fights"]
    summary["roi_even_money"] = summary.apply(lambda row: _roi(row["net_profit_even_money"], row["total_staked"]), axis=1)
    summary["confidence_bucket"] = summary["confidence_bucket"].astype(str)
    return summary[
        [
            "confidence_bucket",
            "fights",
            "correct",
            "accuracy",
            "avg_pick_probability",
            "avg_confidence",
            "total_staked",
            "net_profit_even_money",
            "roi_even_money",
            "expected_profit_even_money",
        ]
    ]


def _yearly_summary(backtest: pd.DataFrame) -> pd.DataFrame:
    yearly = (
        backtest.groupby("year")
        .agg(
            fights=("fight_id", "count"),
            correct=("correct_pick", "sum"),
            total_staked=("stake_if_100_confidence_equals_100", "sum"),
            net_profit_even_money=("net_profit_even_money_sim", "sum"),
            expected_profit_even_money=("expected_profit_even_money_sim", "sum"),
        )
        .reset_index()
    )
    yearly["accuracy"] = yearly["correct"] / yearly["fights"]
    yearly["roi_even_money"] = yearly.apply(lambda row: _roi(row["net_profit_even_money"], row["total_staked"]), axis=1)
    return yearly[["year", "fights", "correct", "accuracy", "total_staked", "net_profit_even_money", "roi_even_money", "expected_profit_even_money"]]


def _threshold_summary(backtest: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for threshold in [0, 5, 10, 15, 20, 25, 30, 40, 50]:
        subset = backtest[backtest["confidence_percent"] >= threshold]
        if subset.empty:
            rows.append((threshold, 0, 0, None, 0.0, 0.0, None, 0.0))
            continue
        total_staked = float(subset["stake_if_100_confidence_equals_100"].sum())
        net_profit = float(subset["net_profit_even_money_sim"].sum())
        rows.append(
            (
                threshold,
                len(subset),
                int(subset["correct_pick"].sum()),
                float(subset["correct_pick"].mean()),
                total_staked,
                net_profit,
                _roi(net_profit, total_staked),
                float(subset["expected_profit_even_money_sim"].sum()),
            )
        )
    return pd.DataFrame(
        rows,
        columns=["min_confidence_percent", "fights", "correct", "accuracy", "total_staked", "net_profit_even_money", "roi_even_money", "expected_profit_even_money"],
    )


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY_FMT = "$#,##0.00;[Red]-$#,##0.00"
PCT_FMT = "0.0%"
NUM_FMT = "0.000"


def _write_dataframe_sheet(workbook: Workbook, sheet_name: str, frame: pd.DataFrame, table_name: str) -> None:
    sheet = workbook.create_sheet(sheet_name)
    for col_idx, column in enumerate(frame.columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=str(column))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="top")
    for row_idx, row in enumerate(frame.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=_excel_value(value))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if not frame.empty:
        table_ref = f"A1:{get_column_letter(len(frame.columns))}{len(frame) + 1}"
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)
    _format_columns(sheet, frame)


def _write_readme_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("README")
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "UFC Predictor Backtest Workbook"
    sheet["A1"].font = Font(bold=True, size=16)
    rows = [
        ("What this is", "Chronological holdout backtest using the same 80/20 split as the model report."),
        ("What this is not", "Not a historical sportsbook odds backtest; no full historical odds file exists locally."),
        ("Odds shown", "Model-implied fair American odds for each fighter and for the model pick."),
        ("P&L shown", "Synthetic even-money +100 simulation with stake = adjusted confidence * $100."),
        ("Confidence formula", "confidence = (pick probability - 0.50) * 2; 50% means $0, 100% means $100."),
    ]
    for row_idx, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row_idx, column=2, value=value)
    sheet["A10"] = "Sheets"
    sheet["A10"].font = Font(bold=True)
    for row_idx, value in enumerate(
        [
            "Summary: key metrics and assumptions.",
            "Fight Backtest: one row per holdout fight.",
            "Confidence Buckets: accuracy and P&L grouped by confidence.",
            "Yearly: yearly performance on holdout fights.",
            "Thresholds: confidence-threshold scenarios.",
            "Charts: visual summary.",
        ],
        start=11,
    ):
        sheet.cell(row=row_idx, column=1, value=value)
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 115


def _write_chart_data_and_charts(workbook: Workbook, backtest: pd.DataFrame, confidence: pd.DataFrame, yearly: pd.DataFrame) -> None:
    chart_data = workbook.create_sheet("Chart Data")
    chart_data.append(["fight_number", "cumulative_profit_even_money_sim", "bankroll_even_money_sim"])
    for idx, row in enumerate(backtest.itertuples(index=False), start=1):
        chart_data.append([idx, float(row.cumulative_profit_even_money_sim), float(row.bankroll_even_money_sim)])

    confidence_start = len(backtest) + 4
    for col_idx, value in enumerate(["confidence_bucket", "accuracy", "roi_even_money", "fights"], start=1):
        chart_data.cell(row=confidence_start, column=col_idx, value=value)
    for offset, row in enumerate(confidence.itertuples(index=False), start=1):
        chart_data.cell(row=confidence_start + offset, column=1, value=row.confidence_bucket)
        chart_data.cell(row=confidence_start + offset, column=2, value=float(row.accuracy))
        chart_data.cell(row=confidence_start + offset, column=3, value=float(row.roi_even_money))
        chart_data.cell(row=confidence_start + offset, column=4, value=int(row.fights))

    year_start = confidence_start + len(confidence) + 4
    for col_idx, value in enumerate(["year", "net_profit_even_money", "accuracy"], start=1):
        chart_data.cell(row=year_start, column=col_idx, value=value)
    for offset, row in enumerate(yearly.itertuples(index=False), start=1):
        chart_data.cell(row=year_start + offset, column=1, value=int(row.year))
        chart_data.cell(row=year_start + offset, column=2, value=float(row.net_profit_even_money))
        chart_data.cell(row=year_start + offset, column=3, value=float(row.accuracy))

    for row_idx in [1, confidence_start, year_start]:
        for col_idx in range(1, 5):
            chart_data.cell(row=row_idx, column=col_idx).fill = HEADER_FILL
            chart_data.cell(row=row_idx, column=col_idx).font = HEADER_FONT
    chart_data.sheet_state = "hidden"

    charts = workbook.create_sheet("Charts")
    charts.sheet_view.showGridLines = False
    charts["A1"] = "Backtest Charts"
    charts["A1"].font = Font(bold=True, size=16)
    charts["A2"] = "P&L charts use synthetic even-money +100 payouts because historical sportsbook odds are not available locally."

    _add_line_chart(
        charts,
        chart_data,
        title="Cumulative Simulated Profit @ Even Money",
        y_title="Profit ($)",
        x_title="Holdout Fight Number",
        data_col=2,
        start_row=1,
        end_row=len(backtest) + 1,
        category_col=1,
        anchor="A4",
    )
    _add_bar_chart(charts, chart_data, "Accuracy By Confidence Bucket", "Accuracy", "Confidence Bucket", 2, confidence_start, confidence_start + len(confidence), 1, "A22")
    _add_bar_chart(charts, chart_data, "ROI By Confidence Bucket @ Even Money", "ROI", "Confidence Bucket", 3, confidence_start, confidence_start + len(confidence), 1, "J22")
    _add_bar_chart(charts, chart_data, "Net Simulated Profit By Year @ Even Money", "Profit ($)", "Year", 2, year_start, year_start + len(yearly), 1, "A40")
    _add_bar_chart(charts, chart_data, "Accuracy By Year", "Accuracy", "Year", 3, year_start, year_start + len(yearly), 1, "J40")


def _add_line_chart(sheet, data_sheet, title: str, y_title: str, x_title: str, data_col: int, start_row: int, end_row: int, category_col: int, anchor: str) -> None:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    chart.add_data(Reference(data_sheet, min_col=data_col, min_row=start_row, max_row=end_row), titles_from_data=True)
    chart.set_categories(Reference(data_sheet, min_col=category_col, min_row=start_row + 1, max_row=end_row))
    chart.height = 9
    chart.width = 22
    sheet.add_chart(chart, anchor)


def _add_bar_chart(sheet, data_sheet, title: str, y_title: str, x_title: str, data_col: int, header_row: int, end_row: int, category_col: int, anchor: str) -> None:
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    chart.add_data(Reference(data_sheet, min_col=data_col, min_row=header_row, max_row=end_row), titles_from_data=True)
    chart.set_categories(Reference(data_sheet, min_col=category_col, min_row=header_row + 1, max_row=end_row))
    chart.height = 9
    chart.width = 18 if anchor.startswith(("A22", "J22")) else 22
    sheet.add_chart(chart, anchor)


def _excel_value(value: object) -> object:
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if pd.isna(value):
        return None
    return value


def _format_columns(sheet, frame: pd.DataFrame) -> None:
    for col_idx, column in enumerate(frame.columns, start=1):
        sample = frame[column].head(250).astype(object)
        values = [str(column)] + [str(value) if pd.notna(value) else "" for value in sample.tolist()]
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(len(value) for value in values) + 2, 42)
        lower = str(column).lower()
        if any(token in lower for token in ["profit", "stake", "bankroll", "loss", "staked"]):
            _format_column(sheet, col_idx, len(frame), MONEY_FMT)
        elif "accuracy" in lower or "roi" in lower or lower in {"confidence", "avg_confidence", "pick_win_probability", "avg_pick_probability", "prob_fighter_a", "prob_fighter_b"}:
            _format_column(sheet, col_idx, len(frame), PCT_FMT)
        elif lower in {"log_loss", "brier_score", "roc_auc"}:
            _format_column(sheet, col_idx, len(frame), NUM_FMT)


def _format_column(sheet, col_idx: int, row_count: int, number_format: str) -> None:
    for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=row_count + 1):
        row[0].number_format = number_format
