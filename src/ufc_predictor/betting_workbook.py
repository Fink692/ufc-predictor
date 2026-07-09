from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .io import ensure_parent


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY_FMT = "$#,##0.00;[Red]-$#,##0.00"
PCT_FMT = "0.0%"
NUM_FMT = "0.000"


def generate_betting_workbook(
    output_path: Path,
    predictions: pd.DataFrame,
    odds_board: pd.DataFrame,
    value_bets: pd.DataFrame,
    fight_recommendations: pd.DataFrame,
    bankroll: float,
    max_confidence_stake: float,
) -> dict[str, object]:
    """Write a polished workbook for the betting report outputs."""
    summary = summarize_betting_report(
        predictions=predictions,
        odds_board=odds_board,
        value_bets=value_bets,
        fight_recommendations=fight_recommendations,
        bankroll=bankroll,
        max_confidence_stake=max_confidence_stake,
    )
    best_lines = _best_lines(value_bets)
    top_matchups = _top_matchup_recommendations(value_bets)

    ensure_parent(output_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_readme_sheet(workbook)
    _write_dataframe_sheet(workbook, "Summary", summary, "BettingSummaryTable")
    _write_dataframe_sheet(workbook, "Fight Recommendations", fight_recommendations, "FightRecommendationsTable")
    _write_dataframe_sheet(workbook, "Value Board", value_bets, "ValueBoardTable")
    _write_dataframe_sheet(workbook, "Best Lines", best_lines, "BestLinesTable")
    _write_dataframe_sheet(workbook, "Top Matchups", top_matchups, "TopMatchupsTable")
    _write_dataframe_sheet(workbook, "Predictions", predictions, "PredictionsTable")
    _write_dataframe_sheet(workbook, "Odds Board", odds_board, "OddsBoardTable")
    _write_chart_data_and_charts(workbook, fight_recommendations, value_bets)
    workbook.save(output_path)

    value_candidate_count = int((value_bets["decision"] == "bet").sum()) if "decision" in value_bets else 0
    return {
        "output_path": str(output_path),
        "fights": int(len(predictions)),
        "sportsbook_lines": int(len(odds_board)),
        "value_candidates": value_candidate_count,
        "confidence_stake_total": _sum_column(fight_recommendations, "confidence_stake"),
        "expected_profit_total": _sum_column(fight_recommendations, "expected_profit"),
    }


def summarize_betting_report(
    predictions: pd.DataFrame,
    odds_board: pd.DataFrame,
    value_bets: pd.DataFrame,
    fight_recommendations: pd.DataFrame,
    bankroll: float,
    max_confidence_stake: float,
) -> pd.DataFrame:
    value_candidates = int((value_bets["decision"] == "bet").sum()) if "decision" in value_bets else 0
    positive_ev_picks = (
        int((fight_recommendations["value_flag"] == "positive_ev").sum())
        if "value_flag" in fight_recommendations
        else 0
    )
    sportsbooks = int(odds_board["sportsbook"].nunique()) if "sportsbook" in odds_board else 0
    total_confidence_stake = _sum_column(fight_recommendations, "confidence_stake")
    expected_profit = _sum_column(fight_recommendations, "expected_profit")
    profit_if_correct = _sum_column(fight_recommendations, "profit_if_correct")
    total_return_if_correct = _sum_column(fight_recommendations, "total_return_if_correct")
    max_loss = _sum_column(fight_recommendations, "max_loss_if_wrong")
    recommended_stake = _sum_column(value_bets[value_bets["decision"] == "bet"], "recommended_stake") if "decision" in value_bets else 0.0
    recommended_expected_profit = _sum_product(
        value_bets[value_bets["decision"] == "bet"] if "decision" in value_bets else value_bets,
        "recommended_stake",
        "expected_roi",
    )

    rows = [
        ("Report type", "Upcoming betting odds report"),
        ("Purpose", "Compare model probabilities against supplied sportsbook moneyline odds"),
        ("Analytics-only notice", "This is not betting advice and does not guarantee profit"),
        ("Fights scored", len(predictions)),
        ("Sportsbook lines ranked", len(odds_board)),
        ("Sportsbooks compared", sportsbooks),
        ("Value candidates", value_candidates),
        ("Positive-EV model picks", positive_ev_picks),
        ("Bankroll used for Kelly sizing", bankroll),
        ("Max confidence stake", max_confidence_stake),
        ("Total confidence stake", total_confidence_stake),
        ("Total expected profit from confidence stakes", expected_profit),
        ("Total profit if every confidence pick wins", profit_if_correct),
        ("Total return if every confidence pick wins", total_return_if_correct),
        ("Total max loss if every confidence pick loses", max_loss),
        ("Total recommended Kelly stake", recommended_stake),
        ("Expected profit from recommended Kelly stake", recommended_expected_profit),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def _best_lines(value_bets: pd.DataFrame) -> pd.DataFrame:
    if value_bets.empty or "best_available_for_fighter" not in value_bets:
        return value_bets.copy()
    columns = [
        "event_date",
        "fighter_a",
        "fighter_b",
        "fighter",
        "sportsbook",
        "american_odds",
        "decimal_odds",
        "model_probability",
        "implied_probability",
        "edge",
        "expected_roi",
        "recommended_stake",
        "potential_profit",
        "decision",
        "risk_label",
    ]
    available = [column for column in columns if column in value_bets]
    return value_bets[value_bets["best_available_for_fighter"]].loc[:, available].reset_index(drop=True)


def _top_matchup_recommendations(value_bets: pd.DataFrame) -> pd.DataFrame:
    if value_bets.empty:
        return value_bets.copy()
    if "best_recommendation_for_matchup" in value_bets and value_bets["best_recommendation_for_matchup"].any():
        return value_bets[value_bets["best_recommendation_for_matchup"]].reset_index(drop=True)
    group_cols = ["event_date", "fighter_a", "fighter_b"]
    if not set(group_cols).issubset(value_bets.columns) or "risk_adjusted_score" not in value_bets:
        return value_bets.head(0).copy()
    indexes = value_bets.groupby(group_cols)["risk_adjusted_score"].idxmax()
    return value_bets.loc[indexes].reset_index(drop=True)


def _write_readme_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("README")
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "UFC Betting Odds Workbook"
    sheet["A1"].font = Font(bold=True, size=16)
    rows = [
        ("What this is", "A workbook version of the betting-report command outputs."),
        ("What this is not", "Not betting advice and not a guarantee of profit."),
        ("Value Board", "Every sportsbook line ranked by model edge, expected ROI, and conservative Kelly sizing."),
        ("Fight Recommendations", "One row per fight using the model pick, best available odds, confidence stake, and expected profit."),
        ("Confidence formula", "confidence = (pick probability - 0.50) * 2; 50% means $0, 100% means the configured max stake."),
        ("Risk framing", "Positive expected value can still lose. Underdog lines can have high variance even when they rank well."),
    ]
    for row_idx, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row_idx, column=2, value=value)
    sheet["A11"] = "Sheets"
    sheet["A11"].font = Font(bold=True)
    for row_idx, value in enumerate(
        [
            "Summary: report-level metrics and sizing assumptions.",
            "Fight Recommendations: model pick, confidence stake, best line, and expected profit per fight.",
            "Value Board: every sportsbook/fighter line ranked for value.",
            "Best Lines: best available line for each fighter across books.",
            "Top Matchups: best risk-adjusted recommendation per matchup.",
            "Predictions: raw model probabilities and confidence.",
            "Odds Board: supplied or fetched sportsbook odds.",
            "Charts: visual summary of stakes, expected profit, and value candidates.",
        ],
        start=12,
    ):
        sheet.cell(row=row_idx, column=1, value=value)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 120


def _write_dataframe_sheet(workbook: Workbook, sheet_name: str, frame: pd.DataFrame, table_name: str) -> None:
    sheet = workbook.create_sheet(sheet_name)
    safe_frame = frame.copy()
    if safe_frame.empty:
        if len(safe_frame.columns) == 0:
            safe_frame = pd.DataFrame({"message": ["No rows available."]})
        else:
            safe_frame = pd.DataFrame([{column: None for column in safe_frame.columns}])
    for col_idx, column in enumerate(safe_frame.columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=str(column))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="top")
    for row_idx, row in enumerate(safe_frame.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=_excel_value(value))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    table_ref = f"A1:{get_column_letter(len(safe_frame.columns))}{len(safe_frame) + 1}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    _format_columns(sheet, safe_frame)


def _write_chart_data_and_charts(workbook: Workbook, fight_recommendations: pd.DataFrame, value_bets: pd.DataFrame) -> None:
    chart_data = workbook.create_sheet("Chart Data")
    chart_data.append(["fight", "confidence_stake", "expected_profit", "edge", "confidence"])
    for row in fight_recommendations.to_dict("records"):
        fight = f"{row.get('fighter_a', '')} vs. {row.get('fighter_b', '')}"
        chart_data.append(
            [
                fight,
                _float_or_zero(row.get("confidence_stake")),
                _float_or_zero(row.get("expected_profit")),
                _float_or_zero(row.get("edge")),
                _float_or_zero(row.get("confidence")),
            ]
        )

    value_start = max(len(fight_recommendations), 1) + 4
    for col_idx, value in enumerate(["line", "expected_roi", "recommended_stake", "risk_adjusted_score"], start=1):
        chart_data.cell(row=value_start, column=col_idx, value=value)
    top_value = value_bets[value_bets["decision"] == "bet"] if "decision" in value_bets else value_bets
    if top_value.empty:
        top_value = value_bets.head(10)
    top_value = top_value.head(12)
    for offset, row in enumerate(top_value.to_dict("records"), start=1):
        line = f"{row.get('fighter', '')} {row.get('sportsbook', '')}"
        chart_data.cell(row=value_start + offset, column=1, value=line)
        chart_data.cell(row=value_start + offset, column=2, value=_float_or_zero(row.get("expected_roi")))
        chart_data.cell(row=value_start + offset, column=3, value=_float_or_zero(row.get("recommended_stake")))
        chart_data.cell(row=value_start + offset, column=4, value=_float_or_zero(row.get("risk_adjusted_score")))

    risk_start = value_start + max(len(top_value), 1) + 4
    chart_data.cell(row=risk_start, column=1, value="risk_label")
    chart_data.cell(row=risk_start, column=2, value="lines")
    risk_counts = (
        value_bets["risk_label"].value_counts().rename_axis("risk_label").reset_index(name="lines")
        if "risk_label" in value_bets
        else pd.DataFrame(columns=["risk_label", "lines"])
    )
    for offset, row in enumerate(risk_counts.itertuples(index=False), start=1):
        chart_data.cell(row=risk_start + offset, column=1, value=row.risk_label)
        chart_data.cell(row=risk_start + offset, column=2, value=int(row.lines))

    for row_idx in [1, value_start, risk_start]:
        for col_idx in range(1, 6):
            cell = chart_data.cell(row=row_idx, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
    chart_data.sheet_state = "hidden"

    charts = workbook.create_sheet("Charts")
    charts.sheet_view.showGridLines = False
    charts["A1"] = "Betting Report Charts"
    charts["A1"].font = Font(bold=True, size=16)
    charts["A2"] = "Charts use the supplied sportsbook lines and model probabilities. They are analytics only, not betting advice."

    fight_end = len(fight_recommendations) + 1
    if fight_end >= 2:
        _add_bar_chart(charts, chart_data, "Confidence Stake By Fight", "Stake ($)", "Fight", 2, 1, fight_end, 1, "A4")
        _add_bar_chart(charts, chart_data, "Expected Profit By Fight", "Expected Profit ($)", "Fight", 3, 1, fight_end, 1, "J4")
        _add_bar_chart(charts, chart_data, "Model Edge By Fight Pick", "Edge", "Fight", 4, 1, fight_end, 1, "A22")
    value_end = value_start + len(top_value)
    if value_end > value_start:
        _add_bar_chart(charts, chart_data, "Top Value Candidate EV Per Dollar", "EV/$", "Line", 2, value_start, value_end, 1, "J22")
        _add_bar_chart(charts, chart_data, "Recommended Kelly Stake", "Stake ($)", "Line", 3, value_start, value_end, 1, "A40")
    risk_end = risk_start + len(risk_counts)
    if risk_end > risk_start:
        _add_bar_chart(charts, chart_data, "Lines By Risk Label", "Lines", "Risk", 2, risk_start, risk_end, 1, "J40")


def _add_bar_chart(
    sheet,
    data_sheet,
    title: str,
    y_title: str,
    x_title: str,
    data_col: int,
    header_row: int,
    end_row: int,
    category_col: int,
    anchor: str,
) -> None:
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    chart.add_data(Reference(data_sheet, min_col=data_col, min_row=header_row, max_row=end_row), titles_from_data=True)
    chart.set_categories(Reference(data_sheet, min_col=category_col, min_row=header_row + 1, max_row=end_row))
    chart.height = 9
    chart.width = 18
    sheet.add_chart(chart, anchor)


def _excel_value(value: object) -> object:
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if pd.isna(value):
        return None
    return value


def _format_columns(sheet, frame: pd.DataFrame) -> None:
    for col_idx, column in enumerate(frame.columns, start=1):
        sample = frame[column].head(250).astype(object)
        values = [str(column)] + [str(value) if pd.notna(value) else "" for value in sample.tolist()]
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(len(value) for value in values) + 2, 44)
        lower = str(column).lower()
        if any(token in lower for token in ["profit", "stake", "bankroll", "loss", "payout", "return"]):
            _format_column(sheet, col_idx, len(frame), MONEY_FMT)
        elif any(token in lower for token in ["probability", "confidence", "edge", "roi", "fraction"]):
            _format_column(sheet, col_idx, len(frame), PCT_FMT)
        elif any(token in lower for token in ["score", "volatility", "decimal_odds"]):
            _format_column(sheet, col_idx, len(frame), NUM_FMT)


def _format_column(sheet, col_idx: int, row_count: int, number_format: str) -> None:
    for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=row_count + 1):
        row[0].number_format = number_format


def _sum_column(frame: pd.DataFrame, column: str) -> float:
    if column not in frame:
        return 0.0
    return float(pd.to_numeric(frame[column], errors="coerce").fillna(0.0).sum())


def _sum_product(frame: pd.DataFrame, left: str, right: str) -> float:
    if left not in frame or right not in frame:
        return 0.0
    left_values = pd.to_numeric(frame[left], errors="coerce").fillna(0.0)
    right_values = pd.to_numeric(frame[right], errors="coerce").fillna(0.0)
    return float((left_values * right_values).sum())


def _float_or_zero(value: object) -> float:
    if value is None or pd.isna(value):
        return 0.0
    return float(value)
