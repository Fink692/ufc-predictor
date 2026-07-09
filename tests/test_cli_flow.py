from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path
import json
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from ufc_predictor.cli import main


FIXTURES = Path(__file__).resolve().parent / "fixtures"


class CliFlowTests(unittest.TestCase):
    def test_train_evaluate_predict_flow(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            features = root / "features.csv"
            model = root / "models" / "ufc_model.joblib"
            train_report = root / "reports" / "train_metrics.json"
            eval_report = root / "reports" / "evaluation.json"
            predictions = root / "reports" / "predictions.csv"
            value_bets = root / "reports" / "value_bets.csv"
            betting_report_csv = root / "reports" / "betting_report.csv"
            fight_recommendations_csv = root / "reports" / "fight_recommendations.csv"
            betting_report_md = root / "reports" / "betting_report.md"

            main(["build-features", "--raw-dir", str(FIXTURES), "--output", str(features)])
            main(["train", "--features", str(features), "--model-path", str(model), "--report", str(train_report)])
            main(
                [
                    "evaluate",
                    "--features",
                    str(features),
                    "--model-path",
                    str(model),
                    "--report",
                    str(eval_report),
                    "--odds",
                    str(FIXTURES / "odds.csv"),
                ]
            )
            main(
                [
                    "predict",
                    "--raw-dir",
                    str(FIXTURES),
                    "--model-path",
                    str(model),
                    "--input",
                    str(FIXTURES / "upcoming_fights.csv"),
                    "--output",
                    str(predictions),
                ]
            )
            main(
                [
                    "rank-odds",
                    "--predictions",
                    str(FIXTURES / "predictions.csv"),
                    "--odds-board",
                    str(FIXTURES / "odds_board.csv"),
                    "--output",
                    str(value_bets),
                    "--bankroll",
                    "1000",
                ]
            )
            main(
                [
                    "betting-report",
                    "--raw-dir",
                    str(FIXTURES),
                    "--model-path",
                    str(model),
                    "--upcoming",
                    str(FIXTURES / "upcoming_fights.csv"),
                    "--odds-board",
                    str(FIXTURES / "odds_board.csv"),
                    "--predictions-output",
                    str(root / "reports" / "report_predictions.csv"),
                    "--output",
                    str(betting_report_csv),
                    "--fight-output",
                    str(fight_recommendations_csv),
                    "--markdown-output",
                    str(betting_report_md),
                    "--max-confidence-stake",
                    "100",
                    "--bankroll",
                    "1000",
                ]
            )

            self.assertTrue(model.exists())
            self.assertTrue(train_report.exists())
            self.assertTrue(eval_report.exists())
            report = json.loads(eval_report.read_text(encoding="utf-8"))
            self.assertIn("walk_forward_event_time", report)
            self.assertIsNotNone(report["walk_forward_event_time"]["metrics"])
            self.assertIn("market_odds_baseline", report)
            self.assertEqual(report["market_odds_baseline"]["coverage_rows"], 12)
            output = pd.read_csv(predictions)
            self.assertEqual(len(output), 2)
            self.assertIn("prob_fighter_a", output.columns)
            self.assertIn("confidence", output.columns)
            self.assertIn("confidence_percent", output.columns)
            self.assertTrue(output["prob_fighter_a"].between(0, 1).all())
            self.assertTrue(output["prob_fighter_b"].between(0, 1).all())
            odds_output = pd.read_csv(value_bets)
            self.assertEqual(len(odds_output), 8)
            self.assertIn("expected_roi", odds_output.columns)
            self.assertTrue((odds_output["decision"] == "bet").any())
            report_output = pd.read_csv(betting_report_csv)
            self.assertEqual(len(report_output), 8)
            self.assertIn("total_payout_if_win", report_output.columns)
            fight_output = pd.read_csv(fight_recommendations_csv)
            self.assertEqual(len(fight_output), 2)
            self.assertIn("confidence_stake", fight_output.columns)
            self.assertIn("profit_if_correct", fight_output.columns)
            self.assertTrue(fight_output["confidence_stake"].between(0, 100).all())
            self.assertTrue(fight_output["profit_if_correct"].ge(0).all())
            self.assertTrue(betting_report_md.exists())
            markdown = betting_report_md.read_text(encoding="utf-8")
            self.assertIn("# UFC Betting Value Report", markdown)
            self.assertIn("# Fight Confidence Bets", markdown)

    def test_betting_report_can_fetch_live_odds_first(self) -> None:
        live_events = [
            {
                "id": "live-1",
                "commence_time": "2023-12-16T23:00:00Z",
                "home_team": "Alpha Adams",
                "away_team": "Ethan Ellis",
                "bookmakers": [
                    {
                        "key": "livebook",
                        "title": "LiveBook",
                        "markets": [
                            {
                                "key": "h2h",
                                "outcomes": [
                                    {"name": "Alpha Adams", "price": 105},
                                    {"name": "Ethan Ellis", "price": -115},
                                ],
                            }
                        ],
                    }
                ],
            },
            {
                "id": "live-2",
                "commence_time": "2023-12-16T23:30:00Z",
                "home_team": "Bruno Blake",
                "away_team": "Felix Fox",
                "bookmakers": [
                    {
                        "key": "livebook",
                        "title": "LiveBook",
                        "markets": [
                            {
                                "key": "h2h",
                                "outcomes": [
                                    {"name": "Bruno Blake", "price": 130},
                                    {"name": "Felix Fox", "price": -115},
                                ],
                            }
                        ],
                    }
                ],
            },
        ]
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            features = root / "features.csv"
            model = root / "models" / "ufc_model.joblib"
            train_report = root / "reports" / "train_metrics.json"
            live_odds = root / "data" / "live_odds.csv"
            predictions = root / "reports" / "predictions.csv"
            value_bets = root / "reports" / "value_bets.csv"
            fight_recommendations = root / "reports" / "fight_recommendations.csv"
            markdown_report = root / "reports" / "betting_report.md"

            main(["build-features", "--raw-dir", str(FIXTURES), "--output", str(features)])
            main(["train", "--features", str(features), "--model-path", str(model), "--report", str(train_report)])

            with patch("ufc_predictor.cli.fetch_odds_api_events", return_value=live_events) as fetch_mock:
                main(
                    [
                        "betting-report",
                        "--raw-dir",
                        str(FIXTURES),
                        "--model-path",
                        str(model),
                        "--upcoming",
                        str(FIXTURES / "upcoming_fights.csv"),
                        "--fetch-live-odds",
                        "--api-key",
                        "test-key",
                        "--odds-board",
                        str(live_odds),
                        "--predictions-output",
                        str(predictions),
                        "--output",
                        str(value_bets),
                        "--fight-output",
                        str(fight_recommendations),
                        "--markdown-output",
                        str(markdown_report),
                    ]
                )

            fetch_mock.assert_called_once()
            live_odds_output = pd.read_csv(live_odds)
            self.assertEqual(len(live_odds_output), 4)
            self.assertEqual(set(live_odds_output["sportsbook"]), {"LiveBook"})
            self.assertTrue(predictions.exists())
            self.assertTrue(value_bets.exists())
            fight_output = pd.read_csv(fight_recommendations)
            self.assertEqual(len(fight_output), 2)
            self.assertEqual(set(fight_output["best_sportsbook"]), {"LiveBook"})
            self.assertIn("# Fight Confidence Bets", markdown_report.read_text(encoding="utf-8"))

    def test_backtest_workbook_command_writes_excel_and_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            features = root / "features.csv"
            workbook = root / "reports" / "backtest.xlsx"
            rows = root / "reports" / "backtest_rows.csv"

            main(["build-features", "--raw-dir", str(FIXTURES), "--output", str(features)])
            main(
                [
                    "backtest-workbook",
                    "--features",
                    str(features),
                    "--output",
                    str(workbook),
                    "--rows-output",
                    str(rows),
                    "--max-confidence-stake",
                    "100",
                ]
            )

            self.assertTrue(workbook.exists())
            self.assertTrue(rows.exists())
            loaded = load_workbook(workbook, read_only=False, data_only=False)
            self.assertIn("Summary", loaded.sheetnames)
            self.assertIn("Fight Backtest", loaded.sheetnames)
            self.assertIn("Charts", loaded.sheetnames)
            self.assertGreaterEqual(len(loaded["Charts"]._charts), 1)
            output_rows = pd.read_csv(rows)
            self.assertGreater(len(output_rows), 0)
            self.assertIn("net_profit_even_money_sim", output_rows.columns)


if __name__ == "__main__":
    unittest.main()
