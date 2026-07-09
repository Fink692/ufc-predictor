from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from ufc_predictor.calibration import calibration_summary, calibration_table
from ufc_predictor.features import build_features_from_raw
from ufc_predictor.historical_odds import backtest_historical_odds, write_historical_odds_workbook
from ufc_predictor.model_compare import compare_model_families


FIXTURES = Path(__file__).resolve().parent / "fixtures"


class AnalyticsTests(unittest.TestCase):
    def test_calibration_table_and_summary(self) -> None:
        table = calibration_table([0, 1, 1, 0], [0.1, 0.7, 0.8, 0.4], bins=2)
        summary = calibration_summary([0, 1, 1, 0], [0.1, 0.7, 0.8, 0.4], bins=2)

        self.assertEqual(len(table), 2)
        self.assertEqual(int(table["row_count"].sum()), 4)
        self.assertIn("expected_calibration_error", summary)
        self.assertGreaterEqual(summary["expected_calibration_error"], 0)

    def test_model_comparison_ranks_multiple_model_families(self) -> None:
        features = build_features_from_raw(FIXTURES)
        comparison = compare_model_families(features)

        self.assertGreaterEqual(len(comparison), 3)
        self.assertIn("regularized_logistic", set(comparison["model"]))
        self.assertIn("expected_calibration_error", comparison.columns)
        self.assertTrue(comparison["log_loss"].notna().all())

    def test_historical_odds_backtest_compares_staking_strategies(self) -> None:
        predictions = pd.read_csv(FIXTURES / "predictions.csv")
        predictions["actual_winner"] = ["Alpha Adams", "Felix Fox"]
        odds_board = pd.read_csv(FIXTURES / "odds_board.csv")

        backtest, summary = backtest_historical_odds(
            predictions,
            odds_board,
            bankroll=1000,
            flat_stake=10,
            max_confidence_stake=100,
            min_edge=0.0,
        )

        self.assertEqual(len(backtest), len(odds_board))
        self.assertIn("kelly_profit", backtest.columns)
        self.assertIn("confidence_bankroll", backtest.columns)
        self.assertEqual(set(summary["strategy"]), {"kelly", "flat", "confidence"})
        self.assertGreater(int(backtest["bet_placed"].sum()), 0)

        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "historical_odds.xlsx"
            write_historical_odds_workbook(workbook_path, backtest, summary)
            loaded = load_workbook(workbook_path, read_only=False, data_only=False)
            self.assertIn("Strategy Summary", loaded.sheetnames)
            self.assertIn("Bet Backtest", loaded.sheetnames)
            self.assertGreaterEqual(len(loaded["Charts"]._charts), 1)


if __name__ == "__main__":
    unittest.main()
